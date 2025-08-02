import os
import pandas as pd
from openpyxl import load_workbook
from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate, Table, TableStyle, Paragraph, Flowable
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# === Custom Flowable to Rotate Text Anticlockwise ===
class RotatedText(Flowable):
    def __init__(self, text, width=35, height=60, fontName="Helvetica-Bold", fontSize=10, wrap_text=False):
        Flowable.__init__(self)
        self.text = text
        self.width = width
        self.height = height
        self.fontName = fontName
        self.fontSize = fontSize
        self.wrap_text = wrap_text

    def wrap(self, availWidth, availHeight):
        return self.width, self.height

    def draw(self):
        self.canv.saveState()
        self.canv.setFont(self.fontName, self.fontSize)
        self.canv.rotate(90)
        self.canv.translate(0, -self.width + 2)
        if self.wrap_text:
            words = self.text.split()
            line_spacing = self.fontSize + 1
            for i, word in enumerate(words):
                self.canv.drawCentredString(self.height / 2, -i * line_spacing, word)
        else:
            self.canv.drawCentredString(self.height / 2, -self.width + 10, self.text)
        self.canv.restoreState()

# ===== CONFIG =====
excel_path = r"C:\Users\Daksh\OneDrive\Desktop\MYFiles\Verification_list\Generated_Verification_By_Center.xlsx"
output_dir = r"C:\Users\Daksh\OneDrive\Desktop\MYFiles\Attendance_sheet\PDF_Centers"
os.makedirs(output_dir, exist_ok=True)

# ===== Styles =====
styles = getSampleStyleSheet()
style_center = ParagraphStyle(name="centered", alignment=TA_CENTER, fontSize=10, leading=10)
style_left = ParagraphStyle(name="lefted", alignment=TA_LEFT, fontSize=10, leading=10)

# Load Excel workbook and get sheet names
wb = load_workbook(excel_path)
sheet_names = wb.sheetnames

# === Process Each Sheet ===
for sheet_name in sheet_names:
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    df.drop(df.columns[[4, 5, 6, 8]], axis=1, inplace=True)
    df.columns = ["Room No", "Seat No", "Dakshana Roll No -- Name", "Gender", "Applied For", "Signature"]

    # === Table Headers ===
    headers = [
        RotatedText("Room No", width=15, height=50, fontSize=10, wrap_text=False),
        RotatedText("Seat No", width=15, height=50, fontSize=10, wrap_text=False),
        Paragraph("<b>Dakshana Roll No -- Name</b>", ParagraphStyle(name="no_wrap", fontSize=10, alignment=TA_CENTER, wordWrap='CJK')),
        RotatedText("Gender", width=15, height=50, fontSize=10, wrap_text=False),
        RotatedText("Applied for (Engg/Med)", width=15, height=80, fontSize=10, wrap_text=True),
        Paragraph("<b>Signature</b>", style_center)
    ]

    # === Table Body ===
    table_data = [headers]
    for row in df.itertuples(index=False):
        table_data.append([
            str(row[0]),
            str(row[1]),
            str(row[2]),
            str(row[3]),
            str(row[4]) if pd.notna(row[4]) else "",
            str(row[5]) if pd.notna(row[5]) else ""
        ])

    # === Table Setup ===
    col_widths = [35, 35, 255, 35, 60, 70]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ALIGN', (0, 1), (1, -1), 'CENTER'),
        ('ALIGN', (3, 1), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
    ]))

    # === PDF Output ===
    output_pdf = os.path.join(output_dir, f"{sheet_name}.pdf")
    doc = BaseDocTemplate(
        output_pdf,
        pagesize=A4,
        leftMargin=0.3 * inch,
        rightMargin=0.3 * inch,
        topMargin=0.3 * inch,
        bottomMargin=0.3 * inch,
    )

    # Reserve space for header box
    header_box_height = 40

    # Create frame for table
    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin,
        doc.width,
        doc.height - header_box_height,
        id='normal'
    )

    # === Header Drawing Function (Canvas) ===
    def draw_page_header(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica-Bold", 14)
        canvas.setFillColor(colors.black)

    # Always center horizontally
        x_center = A4[0] / 2

    # Always fixed from top of the A4 page (use absolute position)
        y_fixed = A4[1] - 40  # 40 points from the top edge

        canvas.drawCentredString(x_center, y_fixed, f"Attendance Sheet: {sheet_name}")
        canvas.restoreState() 
    # Apply the page template
    template = PageTemplate(id='AttendancePage', frames=[frame], onPage=draw_page_header)
    doc.addPageTemplates([template])

    # Build document
    elements = [table]
    doc.build(elements)
    print(f"✅ Saved: {output_pdf}")

print("\n🎯 All PDFs created successfully with centered headers inside boxes on every page.")
