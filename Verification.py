import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

def extract_room_number(text):
    match = re.search(r"Room Number\s*:\s*(\d+)", str(text))
    return int(match.group(1)) if match else None

def generate_verification_layout_by_center(seat_file_path, output_path, gender_file_path):
    seat_sheets = pd.read_excel(seat_file_path, sheet_name=None)
    gender_df = pd.read_excel(gender_file_path)
    gender_df.columns = gender_df.columns.str.strip()

    gender_df["Dakshana Roll No -- Name"] = (
        gender_df["Dakshana Roll No -- Name"]
        .astype(str)
        .str.replace("  ", " ")
        .str.replace(" --", "--")
        .str.replace("-- ", "--")
        .str.strip()
    )
    gender_df["M/F"] = gender_df["M/F"].astype(str).str.strip()
    gender_map = dict(zip(
        gender_df["Dakshana Roll No -- Name"],
        gender_df["M/F"].apply(lambda x: 'M' if x.lower() == 'male' else 'F')
    ))

    center_data = {}

    for sheet_name, df in seat_sheets.items():
        center_name = sheet_name.strip()
        center_data.setdefault(center_name, [])
        room_no = None
        seat_counter = 1

        for i, row in df.iterrows():
            for cell in row:
                if isinstance(cell, str) and "Room Number" in cell:
                    room_match = extract_room_number(cell)
                    if room_match is not None:
                        room_no = room_match
                        seat_counter = 1

            if room_no is None or i < 4:
                continue

            for col in range(2, len(row)):
                val = row[col]
                if isinstance(val, str) and "--" in val:
                    parts = val.split("--")
                    if len(parts) == 2:
                        roll = parts[0].strip()
                        name = parts[1].strip()
                        full_name = f"{roll}--{name}".replace("  ", " ").strip()
                        gender = gender_map.get(full_name, "")

                        center_data[center_name].append({
                            "Room No": room_no,
                            "Seat No": seat_counter,
                            "Dakshana Roll No -- Name": f"{roll} -- {name}",
                            "Gender": gender,
                            "Aadhar No (Last 4 digit)": "",
                            "Admit card verified (Yes/No)": "",
                            "Roll Number verified in OMR (Yes/No)": "",
                            "Applied for (Engg/Med)": "",
                            "Paste student passport size photograph": "",
                            "Remark": ""
                        })
                        seat_counter += 1

    if os.path.exists(output_path):
        os.remove(output_path)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for center_name, data in center_data.items():
            df = pd.DataFrame(data)
            df.to_excel(writer, index=False, sheet_name=center_name[:31])

    apply_borders_to_excel(output_path)
    format_excel(output_path)
    print(f"✅ Verification workbook created at: {output_path}")

def apply_borders_to_excel(excel_path):
    wb = load_workbook(excel_path)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

    wb.save(excel_path)
    print(f"✅ Borders applied to the Excel file: {excel_path}")

def format_excel(output_path):
    wb = load_workbook(output_path)
    horizontal_headers = ["Dakshana Roll No -- Name", "Remark"]

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header_row = ws[1]

        column_dimensions = {
            "Room No": 3,
            "Seat No": 3,
            "Dakshana Roll No -- Name": 25,
            "Gender": 3,
            "Aadhar No (Last 4 digit)": 6,
            "Admit card verified (Yes/No)": 8,
            "Roll Number verified in OMR (Yes/No)": 8,
            "Applied for (Engg/Med)": 8,
            "Paste student passport size photograph": 16,
            "Remark": 12
        }

        for col_num, cell in enumerate(header_row, 1):
            header = str(cell.value).strip()
            col_letter = get_column_letter(col_num)

            cell.alignment = Alignment(
                textRotation=0 if header in horizontal_headers else 90,
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            cell.font = Font(bold=True)
            ws.column_dimensions[col_letter].width = column_dimensions.get(header, 15)

        ws.row_dimensions[1].height = 100

        for row in ws.iter_rows(min_row=2):
            ws.row_dimensions[row[0].row].height = 100
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    wb.save(output_path)

# === File Paths ===
seat_file = r"C:\Users\Daksh\OneDrive\Desktop\MYFiles\Seat_allotment\Seat_Allotment_Result_Final.xlsx"
output_file = r"C:\Users\Daksh\OneDrive\Desktop\MYFiles\Verification_list\Generated_Verification_By_Center.xlsx"
gender_file = r"C:\Users\Daksh\OneDrive\Desktop\MYFiles\Data_For_all\Data_Student.xlsx"

# === Run ===
generate_verification_layout_by_center(seat_file, output_file, gender_file)
